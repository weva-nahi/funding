"""Opportunity services — write operations."""

import datetime
import logging

from django.db import IntegrityError, transaction
from django.utils import timezone

from common.exceptions import ConflictError, InvalidFileError
from common.utils.hashing import generate_opportunity_hash

from .models import FundingOpportunity, SavedOpportunity

logger = logging.getLogger(__name__)

COMPLETENESS_FIELDS = [
    "title",
    "description",
    "deadline",
    "amount",
    "country",
    "eligibility_criteria",
    "funding_type",
    "sector",
    "url",
]

# Valid source choices matching the model
VALID_SOURCES = {"GEF", "GCF", "OECD", "CLIMATE_FUND", "WORLD_BANK", "AFD", "EU"}

# Common aliases people type in Excel that should map to valid choices
SOURCE_ALIASES = {
    "WORLD BANK": "WORLD_BANK",
    "WORLD-BANK": "WORLD_BANK",
    "CLIMATE FUND": "CLIMATE_FUND",
    "CLIMATE-FUND": "CLIMATE_FUND",
    "GREEN CLIMATE FUND": "GCF",
    "GLOBAL ENVIRONMENT FACILITY": "GEF",
    "AGENCE FRANCAISE DE DEVELOPPEMENT": "AFD",
    "EUROPEAN UNION": "EU",
}


def _normalize_source(raw: str) -> str:
    """Map a raw source string to a valid model choice, defaulting to OECD."""
    if not raw:
        return "OECD"
    upper = str(raw).strip().upper()
    if upper in VALID_SOURCES:
        return upper
    if upper in SOURCE_ALIASES:
        return SOURCE_ALIASES[upper]
    return "OECD"


def create_opportunity(*, created_by=None, **kwargs) -> FundingOpportunity:
    if "hash" not in kwargs:
        kwargs["hash"] = generate_opportunity_hash(
            kwargs.get("title", ""),
            kwargs.get("url", ""),
            kwargs.get("amount", ""),
        )
    if not kwargs.get("completeness_score"):
        kwargs["completeness_score"] = _calculate_completeness(kwargs)
    if created_by is not None:
        kwargs["created_by"] = created_by
    if "source" in kwargs:
        kwargs["source"] = _normalize_source(kwargs["source"])

    if FundingOpportunity.objects.filter(hash=kwargs["hash"]).exists():
        raise ConflictError(
            "An opportunity with the same title, URL and amount already exists."
        )

    try:
        with transaction.atomic():
            return FundingOpportunity.objects.create(**kwargs)
    except IntegrityError as exc:
        raise ConflictError(
            "An opportunity with the same title, URL and amount already exists."
        ) from exc


def update_opportunity(
    *, opportunity: FundingOpportunity, **kwargs
) -> FundingOpportunity:
    for field, value in kwargs.items():
        if hasattr(opportunity, field):
            setattr(opportunity, field, value)
    opportunity.completeness_score = _calculate_completeness_from_instance(opportunity)
    opportunity.save()
    return opportunity


def publish_opportunity(
    *, opportunity: FundingOpportunity, user=None
) -> FundingOpportunity:
    opportunity.status = "published"
    opportunity.published_at = timezone.now()
    opportunity.save(update_fields=["status", "updated_at", "published_at"])
    return opportunity


def bulk_publish_opportunities(*, ids: list, user=None) -> dict:
    """Publish many opportunities at once. Returns counts."""
    now = timezone.now()
    qs = FundingOpportunity.objects.filter(id__in=ids).exclude(status="published")
    published = 0
    for opp in qs:
        opp.status = "published"
        opp.published_at = now
        opp.save(update_fields=["status", "published_at", "updated_at"])
        published += 1
    return {"published": published, "requested": len(ids)}


def archive_opportunity(*, opportunity: FundingOpportunity) -> FundingOpportunity:
    opportunity.status = "archived"
    opportunity.save(update_fields=["status", "updated_at"])
    return opportunity


# ─── Save-for-later ───────────────────────────────────────────────────────────

def save_opportunity(*, user, opportunity: FundingOpportunity) -> SavedOpportunity:
    obj, _created = SavedOpportunity.objects.get_or_create(
        user=user, opportunity=opportunity
    )
    return obj


def unsave_opportunity(*, user, opportunity_id: int) -> None:
    SavedOpportunity.objects.filter(user=user, opportunity_id=opportunity_id).delete()


def import_opportunities_from_xlsx(*, file_obj, created_by=None) -> dict:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise InvalidFileError(f"Could not read the Excel file: {exc}") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        raw_header = next(rows)
    except StopIteration:
        raise InvalidFileError("The Excel file is empty.")

    header = [
        str(h).strip().lower() if h is not None else "" for h in raw_header
    ]

    field_map = {
        "title": "title",
        "source": "source",
        "description": "description",
        "country": "country",
        "amount": "amount",
        "currency": "currency",
        "deadline": "deadline",
        "funding_type": "funding_type",
        "sector": "sector",
        "url": "url",
        "eligibility_criteria": "eligibility_criteria",
        "required_documents": "required_documents",
    }
    col_index = {field_map[h]: i for i, h in enumerate(header) if h in field_map}

    if "title" not in col_index:
        raise InvalidFileError(
            "The Excel file must contain at least a 'title' column."
        )

    created = skipped = 0

    for row in rows:
        if row is None:
            continue

        data = {}
        for field, idx in col_index.items():
            value = row[idx] if idx < len(row) else None
            if value is None:
                continue

            if isinstance(value, (datetime.date, datetime.datetime)):
                value = value.strftime("%Y-%m-%d")
            else:
                value = str(value).strip()

            if value:
                data[field] = value

        if not data.get("title"):
            skipped += 1
            continue

        data["source"] = _normalize_source(data.get("source", ""))
        data.setdefault("currency", "USD")

        if "amount" in data:
            try:
                data["amount"] = float(str(data["amount"]).replace(",", ""))
            except (ValueError, TypeError):
                data.pop("amount")

        h = generate_opportunity_hash(
            data.get("title", ""),
            data.get("url", ""),
            data.get("amount", ""),
        )
        if FundingOpportunity.objects.filter(hash=h).exists():
            skipped += 1
            continue

        try:
            create_opportunity(created_by=created_by, status="draft", hash=h, **data)
            created += 1
        except Exception as exc:  # noqa: BLE001
            logger.warning("Skipping row due to error: %s | data=%s", exc, data)
            skipped += 1

    wb.close()
    return {"created": created, "skipped": skipped}


def _calculate_completeness(data: dict) -> int:
    completed = sum(1 for f in COMPLETENESS_FIELDS if data.get(f))
    return int((completed / len(COMPLETENESS_FIELDS)) * 100)


def _calculate_completeness_from_instance(opp: FundingOpportunity) -> int:
    completed = sum(1 for f in COMPLETENESS_FIELDS if getattr(opp, f, None))
    return int((completed / len(COMPLETENESS_FIELDS)) * 100)